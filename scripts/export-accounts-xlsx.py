"""Export seed accounts to docs/accounts.xlsx."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
SEED = ROOT / "infra" / "local" / "seed" / "demo-students.json"
OUT = ROOT / "docs" / "accounts.xlsx"

SERVICE_IDS = {
    "admin-platform",
    "student-temp-local",
    "demo-guardian",
    "demo-coach",
    "demo-renter",
    "student-synthetic-adult",
}


def role_of(account: dict) -> str:
    roles = account.get("roles") or []
    if roles:
        return ", ".join(roles)
    return str(account.get("role") or "student")


def style_header(ws, headers: list[str]) -> None:
    fill = PatternFill("solid", fgColor="1F2937")
    font = Font(bold=True, color="FFFFFF")
    thin = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    for col, title in enumerate(headers, 1):
        cell = ws.cell(1, col, title)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin
    ws._thin = thin  # type: ignore[attr-defined]


def main() -> None:
    data = json.loads(SEED.read_text(encoding="utf-8"))
    accounts = data["accounts"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Учётки"

    headers = ["Имя", "Роль", "Логин", "Пароль", "Student ID", "Character ID"]
    style_header(ws, headers)
    thin = ws._thin  # type: ignore[attr-defined]
    alt = PatternFill("solid", fgColor="F3F4F6")

    for row_idx, account in enumerate(accounts, 2):
        values = [
            account.get("name", ""),
            role_of(account),
            account.get("login", ""),
            account.get("password", ""),
            account.get("studentId", ""),
            account.get("characterId", ""),
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row_idx, col, value)
            cell.border = thin
            cell.alignment = Alignment(vertical="center")
            if row_idx % 2 == 0:
                cell.fill = alt

    for i, width in enumerate([28, 22, 42, 16, 28, 28], 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.auto_filter.ref = f"A1:F{len(accounts) + 1}"
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Служебные")
    style_header(ws2, ["Имя", "Роль", "Логин", "Пароль"])
    thin2 = ws2._thin  # type: ignore[attr-defined]
    row = 2
    for account in accounts:
        if account.get("studentId") not in SERVICE_IDS:
            continue
        values = [
            account.get("name", ""),
            role_of(account),
            account.get("login", ""),
            account.get("password", ""),
        ]
        for col, value in enumerate(values, 1):
            cell = ws2.cell(row, col, value)
            cell.border = thin2
        row += 1
    for i, width in enumerate([28, 22, 36, 14], 1):
        ws2.column_dimensions[get_column_letter(i)].width = width
    ws2.freeze_panes = "A2"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT} ({len(accounts)} accounts)")


if __name__ == "__main__":
    main()
