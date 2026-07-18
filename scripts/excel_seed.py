#!/usr/bin/env python3
"""Seed demo accounts from Мастер Меча.xlsx (school mastery ledger)."""
from __future__ import annotations

import hashlib
import json
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess, sys

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

ROOT = Path(__file__).resolve().parents[1]
OLD = ROOT / "Master of the Sword module" / "Old"
WEAPONS = {
    "spada a uno mano": "spada_a_uno_mano",
    "due spade": "due_spade",
    "spada e scudo": "spada_e_scudo",
    "spada a due mani": "spada_a_due_mani",
    "spadone": "spadone",
    "acia & alabarda": "acia_alabarda",
    "spiedo & partesana": "spiedo_partesana",
    "spiedo & scudo": "spiedo_e_scudo",
    "spiedo e scudo": "spiedo_e_scudo",
}
REPL = {
    "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "e",
    "ж": "zh", "з": "z", "и": "i", "й": "y", "к": "k", "л": "l", "м": "m",
    "н": "n", "о": "o", "п": "p", "р": "r", "с": "s", "т": "t", "у": "u",
    "ф": "f", "х": "h", "ц": "ts", "ч": "ch", "ш": "sh", "щ": "sch",
    "ъ": "", "ы": "y", "ь": "", "э": "e", "ю": "yu", "я": "ya",
}


def slugify(name: str) -> str:
    s = name.lower().strip()
    out = []
    for ch in s:
        if ch in REPL:
            out.append(REPL[ch])
        elif "a" <= ch <= "z" or "0" <= ch <= "9":
            out.append(ch)
        else:
            out.append("-")
    slug = re.sub(r"-+", "-", "".join(out)).strip("-") or "student"
    return slug


def title_slug(slug: str) -> str:
    return "".join(p.capitalize() for p in slug.split("-"))


def points_to_units(points: float) -> int:
    return int(round(points * 10000))


def rank_from_units(units: int) -> int:
    if units <= 0:
        return 0
    return min(10, units // 100000)


def main() -> None:
    xlsx = max(OLD.glob("*.xlsx"), key=lambda p: p.stat().st_size)
    raw = xlsx.read_bytes()
    digest = hashlib.sha256(raw).hexdigest()
    as_of = date.today()
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    # Jan.. sheets in order
    month_idx = as_of.month - 1
    sheet_name = wb.sheetnames[month_idx] if month_idx < len(wb.sheetnames) else wb.sheetnames[-1]
    ws = wb[sheet_name]

    # day -> column (1-based). Header row typically 3 with datetime cells.
    day_cols: dict[int, int] = {}
    for r in range(1, 8):
        for c in range(1, 40):
            v = ws.cell(r, c).value
            if isinstance(v, datetime):
                if v.month == as_of.month:
                    day_cols[v.day] = c
            elif isinstance(v, date) and not isinstance(v, datetime):
                if v.month == as_of.month:
                    day_cols[v.day] = c
        if len(day_cols) >= 20:
            break
    if not day_cols:
        # fallback positional: col 3 = day 1
        for d in range(1, 32):
            day_cols[d] = 2 + d

    accounts = []
    r = 1
    max_r = ws.max_row or 1
    while r <= max_r:
        a = ws.cell(r, 1).value
        if a is None:
            r += 1
            continue
        name = str(a).strip()
        if name.lower() in WEAPONS:
            r += 1
            continue
        has_w = False
        for rr in range(r + 1, min(r + 5, max_r + 1)):
            na = ws.cell(rr, 1).value
            if na and str(na).strip().lower() in WEAPONS:
                has_w = True
                break
        if not has_w:
            r += 1
            continue

        points: dict[str, float] = {}
        rr = r + 1
        # skip blanks / date header rows until first weapon or next student
        while rr <= max_r:
            wa = ws.cell(rr, 1).value
            if wa is None:
                rr += 1
                continue
            wname = str(wa).strip()
            if wname.lower() in WEAPONS:
                break
            # next student block
            if any(
                ws.cell(x, 1).value and str(ws.cell(x, 1).value).strip().lower() in WEAPONS
                for x in range(rr + 1, min(rr + 5, max_r + 1))
            ):
                break
            rr += 1

        while rr <= max_r:
            wa = ws.cell(rr, 1).value
            if wa is None:
                # companion row or gap — peek for next weapon
                nxt = None
                for x in range(rr + 1, min(rr + 3, max_r + 1)):
                    nv = ws.cell(x, 1).value
                    if nv and str(nv).strip().lower() in WEAPONS:
                        nxt = x
                        break
                    if nv and str(nv).strip().lower() not in WEAPONS:
                        nxt = -1
                        break
                if nxt and nxt > 0:
                    rr = nxt
                    continue
                break
            wname = str(wa).strip()
            if wname.lower() not in WEAPONS:
                break
            opening = ws.cell(rr, 2).value
            try:
                running = float(opening or 0)
            except (TypeError, ValueError):
                running = 0.0
            for d in range(1, as_of.day + 1):
                c = day_cols.get(d)
                if not c:
                    continue
                dv = ws.cell(rr, c).value
                if dv is None:
                    continue
                try:
                    running += float(dv)
                except (TypeError, ValueError):
                    pass
            points[wname] = running
            rr += 2

        slug = slugify(name)
        units = {WEAPONS[k.lower()]: points_to_units(v) for k, v in points.items() if k.lower() in WEAPONS}
        ranks = {k: rank_from_units(v) for k, v in units.items()}
        accounts.append(
            {
                "studentId": f"student-{slug}",
                "name": name,
                "login": f"demo.{slug}@masterofsword.local",
                "password": f"MoS-Demo-{title_slug(slug)}-2026!",
                "characterId": f"char-{slug}",
                "masteryPointsAsOf": points,
                "masteryUnits": units,
                "ranks": ranks,
            }
        )
        r = max(rr, r + 1)

    accounts.append(
        {
            "studentId": "student-synthetic-adult",
            "name": "Synthetic Adult",
            "login": "demo.adult@masterofsword.local",
            "password": "MoS-Demo-Adult-2026!",
            "characterId": "char-synthetic-adult",
            "masteryPointsAsOf": {},
            "masteryUnits": {},
            "ranks": {},
        }
    )

    out = {
        "source": xlsx.name,
        "sourceHash": digest,
        "asOf": as_of.isoformat(),
        "sheet": sheet_name,
        "count": len(accounts),
        "accounts": accounts,
    }
    seed_dir = ROOT / "infra" / "local" / "seed"
    seed_dir.mkdir(parents=True, exist_ok=True)
    (seed_dir / "demo-students.json").write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    (ROOT / "infra" / "local" / "import-state.json").write_text(json.dumps({"hash": digest}), encoding="utf-8")

    md = [
        "# Demo Accounts (local/staging only)\n",
        f"\nSource: `{xlsx.name}` hash `{digest[:12]}` as-of `{as_of.isoformat()}`\n",
        "\n| Name | Login | Password |\n|---|---|---|\n",
    ]
    for a in accounts:
        md.append(f"| {a['name']} | `{a['login']}` | `{a['password']}` |\n")
    (ROOT / "docs" / "demo-accounts.md").write_text("".join(md), encoding="utf-8")

    nonempty = sum(1 for a in accounts if a["masteryUnits"])
    print(f"seeded {len(accounts)} accounts ({nonempty} with mastery) sheet={sheet_name} asOf={as_of}")


if __name__ == "__main__":
    main()
